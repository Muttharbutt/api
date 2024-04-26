from fastapi import FastAPI
import requests
from typing import List, Dict
import openpyxl
import emoji
import json
from unidecode import unidecode
app = FastAPI()
prename=""
checklang=""
def remove_accents(text):
    return unidecode(text)
def remove_emoji(string):
    return emoji.replace_emoji(string, replace='')
def process_prénom(value):
    if value and value.strip():  
        return ','.join(char for char in value if char != ' ')
    return ''
def checklangauge(value):
    global checklang
    products=value.upper()
    products_list = products.split(',')
    first_product = products_list[0]
    
    wb = openpyxl.load_workbook("./list.xlsx")
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        product = row[4]
        product_upper = product.upper()
        if product_upper==first_product:
            checklang='fr'
        product = row[7]
        product_upper = product.upper()
        if product_upper==first_product:
            checklang='en'
        product = row[8]
        product_upper = product.upper()
        if product_upper==first_product:
            checklang='de'
def process_value(key, value, prénom_encountered):
    global prename
    if key == "Prénom":
        if prénom_encountered:
            prename=value
            return '' + process_prénom(value)
        else:
            return process_prénom(value)
    if key == "Child Name":
        if prénom_encountered:
            prename=value
            return '' + process_prénom(value)
        else:
            return process_prénom(value) 
    if key == "Vorname":
        if prénom_encountered:
            prename=value
            return '' + process_prénom(value)
        else:
            return process_prénom(value)  
    else:
        return value
def compare_products(filename, products):
    # Convert products to uppercase
    products_list = products.split(',')
    products_list = [product.strip().replace("ß", "B") for product in products_list] 
    products_upper = [product.upper() for product in products_list]
    products_upper = [product.replace("&AMP;", "&").replace("Ö", "O").replace("Ä", "A").replace("Ü", "U") for product in products_upper]
    print(products_upper)
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    matched_values = []
    for row in sheet.iter_rows(values_only=True):
        if checklang=="fr":
            product = row[4]
        if checklang=="de":
            product = row[8]
        if checklang=="en":
            product = row[7]
        product_upper = product.upper()
        for pro in products_upper:
            if pro==product_upper:
                matched_values.append(row)
    wb.close()
    print(matched_values)
    return matched_values
@app.post("/process_data")
async def process_data(data: List[Dict]):
    modified_data = []
    totalwieght=0
    for d in data:
        
        d["StoreKey"] = "KITIMIMI"
        d["OrderType"] = "Sell" 
        full_name = d["shipping_full_name"]  
        first_name, last_name = full_name.split(maxsplit=1) 
        
        static_fields_to_remove = [key for key in d.keys() if key.startswith("static_field")]
        # for key in list(d.keys()):
        #     if key.startswith("customer_note"):
        #         new_key = "CheckoutMessage" + key[len("customer_note"):]
        #         d[new_key] = d.pop(key)
        for key in list(d.keys()):
            if key.startswith("order_number"):
                new_key = "OrderKey" + key[len("order_number"):]
                d[new_key] = d.pop(key)
        for key in list(d.keys()):
            if key.startswith("order_date"):
                new_key = "OrderDate" + key[len("order_date"):]
                d[new_key] = d.pop(key)
        for key in static_fields_to_remove:
            del d[key]
        prénom_encountered = False
        products = d.get('products', [])
        result = []
        target_keys = ["Prénom", "Child Name", "Vorname"]
        formatted_data = {}
        for product in products:
            final_result=''
            processed_values = []
            for key, value in product.items():
                value = remove_accents(value)
                if key not in target_keys and value is not None and value != '':
                    checklangauge(value)
                    break
            for key, value in product.items():
                value = remove_accents(value)
                if value is not None and value != '':
                    value_without_emojis = remove_emoji(value)
                    if key == "Prénom":
                        prénom_encountered = True
                        if "CheckoutMessage" not in d:
                            d["CheckoutMessage"] = value
                        else:
                            d["CheckoutMessage"] =d["CheckoutMessage"] +" ,"+ value   
                    processed_value = process_value(key, value_without_emojis, prénom_encountered)
                    if r"\u00é" in processed_value:
                        processed_value = processed_value.replace(r'\u00é', 'e')
                    if r"é" in processed_value:
                        processed_value = processed_value.replace(r'é', 'e')
                    if r"ç" in processed_value:
                        processed_value = processed_value.replace(r'ç', 'c')
                    processed_values.append(processed_value)
            result.append(','.join(processed_values))
            final_result = ','.join(result)
            checking=compare_products("./list.xlsx",final_result)
            result=[]
            for item in checking:
                formatted_item = "{:.0f}".format(item[0])
                sku = formatted_item + str(item[1]) + str(item[2])  
                quantity = 1  # Quantity is the sixth element
                weight =  item[5]   # Weight is the seventh element, if None, make it empty
                ProductName=item[4]
                totalwieght=weight+totalwieght
                if sku not in formatted_data:
                    formatted_data[sku] = {
                        "SKU": sku,
                        "Quantity": quantity,
                        "ProductName": ProductName,  # Empty for now, will be filled later
                        "CN23Category": "",
                        "PriceExclTax": "",
                        "Weight": weight,
                        "EANCode": "",
                        "VariationID": "",
                        "VAT": ""
                    }
                else:
                    formatted_data[sku]["Quantity"] += quantity
        # Convert the dictionary to a list of formatted dictionaries
        formatted_list = list(formatted_data.values())
        del d['products']
        d["products"] = formatted_list
        d["Delivery"] = {
        "Recipient":{
                    "RecipLanguageCode":"FR",
                    "RecipCompanyName":"",
                    "RecipFirstName":remove_accents(first_name),
                    "RecipLastName":remove_accents(last_name),
                    "RecipAdr2":"",
                    "RecipAdr1":remove_accents(d["shipping_address_2"]),
                    "RecipAdr0":remove_accents(d["shipping_address_1"]),
                    "RecipZipCode":remove_accents(d["shipping_postcode"]),
                    "RecipCity":remove_accents(d["shipping_city"]),
                    "RecipCountryCode":remove_accents(d["shipping_country"]),
                    "RecipCountryLib":remove_accents(d["shipping_country_full"]),
                    "Recipemail":remove_accents(d["billing_email"]),
                    "RecipMobileNumber":"",
                    "RecipPhoneNumber":d["shipping_phone"],
                    "DeliveryRelayCountry":"",
                    "DeliveryRelayNumber":""
                },
                "Parcel":{
                    "ShippingServiceKey":"",
                    "ShippingProductCode":"",
                    "InsurranceYN":"",
                    "InsurranceCurrency":"",
                    "ParcelValueCurrency":"",
                    "ParcelWeight":totalwieght,
                    "WeightUnit":"g",
                    "DeliveryInstructions1":prename
                },
                "Sender":{
                    "SenderLanguageCode":"FR",
                    "SenderAdr2":"51 Rue de Toufflers",
                    "SenderAdr3":"POUR KITIMIMI",
                    "Sendercity":" Lys-lez-Lannoy",
                    "SenderzipCode":"59390",
                    "SendercompanyName":" Stock logistic",
                    "SendercountryCode":"FR",
                    "SendercountryLib":"FRANCE",
                    "SenderphoneNumber":"0033320200909",
                    "Senderemail":"hello@kitimimi.com"
                }

    }   
        totalwieght=0
        if 'billing_email' in d:
            del d['billing_email']
        if 'order_number' in d:
            del d['order_number']
        if 'shipping_address_1' in d:
            del d['shipping_address_1']
        if 'shipping_address_2' in d:
            del d['shipping_address_2']
        if 'shipping_city' in d:
            del d['shipping_city']
        if 'shipping_country' in d:
            del d['shipping_country']
        if 'shipping_full_name' in d:
            del d['shipping_full_name']
        if 'shipping_phone' in d:
            del d['shipping_phone']
        if 'shipping_postcode' in d:
            del d['shipping_postcode']
        if 'shipping_country_full' in d:
            del d['shipping_country_full']
        if 'order_total_inc_refund' in d:
            del d['order_total_inc_refund']
        if 'order_total_tax_minus_refund' in d:
            del d['order_total_tax_minus_refund']
        if 'payment_method' in d:
            del d['payment_method']
        modified_data.append(d)
    # Construct response data
    response_data ={
        "Request": 
    {
        "Orders": modified_data
    }}
    file_name = "response_data.json"

    with open(file_name, "w") as json_file:
        json.dump(response_data, json_file)

    # headers = {
    #     "Token": "KEy5YrFM3EieHYc+CSoFTZlFBtVonvat" 
    # }x    
    # api_url = "https://sl.atomicseller.com/Api/Order/CreateOrders"
    # response = requests.post(
    #     url=api_url,
    #     json=response_data,
    #     headers=headers
    # )

    # if response.status_code == 200:
    #     print("Response sent successfully.")
    # else:
    #     print("Failed to send response. Status code:", response.status_code)